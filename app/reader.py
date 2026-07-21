"""Чтение прайса в «сетку» — простую таблицу list[list[str]].

Главная сложность реальных прайсов — объединённые ячейки: в Excel значение
хранится только в левой верхней ячейке диапазона, а остальные пусты. Если читать
наивно, шапки и подзаголовки «разъезжаются». Поэтому при чтении xlsx мы
разворачиваем объединения: копируем значение во все ячейки диапазона.

CSV читаем с автоопределением разделителя (запятая/точка с запятой/таб).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook

# Тип «сетка»: строки таблицы, в каждой — ячейки, приведённые к строкам.
Grid = list[list[str]]


class ReadError(RuntimeError):
    """Файл не удалось прочитать (не тот формат, битый файл, нет листа)."""


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # 12.0 из Excel показываем как "12", а не "12.0".
        return str(int(value))
    return str(value).strip()


def _read_xlsx(path: Path, sheet: Optional[Union[str, int]]) -> Grid:
    # data_only=True — берём посчитанные значения формул, а не сам текст формулы.
    workbook = load_workbook(path, data_only=True, read_only=False)

    if sheet is None:
        worksheet = workbook.active
    elif isinstance(sheet, int):
        worksheet = workbook.worksheets[sheet]
    else:
        if sheet not in workbook.sheetnames:
            raise ReadError(
                f"Лист «{sheet}» не найден. Доступны: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet]

    # 1. Читаем как есть.
    grid: Grid = [
        [_cell_to_str(cell) for cell in row]
        for row in worksheet.iter_rows(values_only=True)
    ]

    # 2. Разворачиваем объединённые ячейки: значение из левого-верхнего угла
    #    диапазона копируем во все ячейки этого диапазона.
    for merged in list(worksheet.merged_cells.ranges):
        top_value = _cell_to_str(worksheet.cell(merged.min_row, merged.min_col).value)
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                gr, gc = r - 1, c - 1  # openpyxl 1-based → сетка 0-based
                if gr < len(grid) and gc < len(grid[gr]):
                    grid[gr][gc] = top_value

    return grid


def _read_csv(path: Path) -> Grid:
    # Читаем сырой текст и пытаемся угадать разделитель по первым строкам.
    text = path.read_text(encoding="utf-8-sig")  # utf-8-sig убирает BOM из Excel-CSV
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    return [[cell.strip() for cell in row] for row in reader]


def read_grid(path: Union[str, Path], sheet: Optional[Union[str, int]] = None) -> Grid:
    """Прочитать прайс (xlsx или csv) в сетку строк."""
    path = Path(path)
    if not path.exists():
        raise ReadError(f"Файл не найден: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        grid = _read_xlsx(path, sheet)
    elif suffix in {".csv", ".txt"}:
        grid = _read_csv(path)
    else:
        raise ReadError(f"Формат «{suffix}» не поддерживается. Нужен .xlsx или .csv")

    if not grid:
        raise ReadError("Файл пуст — нет ни одной строки.")
    return grid
