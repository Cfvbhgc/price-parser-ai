"""Формирование красивого Excel-отчёта через openpyxl.

Оформление:
  - строка-заголовок отчёта (объединённая) с названием;
  - тёмная шапка таблицы с белым жирным шрифтом;
  - чередование строк (зебра) для читаемости;
  - тонкие границы и включённая сетка;
  - автоширина колонок, автофильтр, закреплённая шапка;
  - формат цены как числа с разделителем тысяч.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Product

COLUMNS = ["Название", "Артикул", "Цена", "Категория"]

# Палитра.
HEADER_FILL = PatternFill("solid", fgColor="2E3A4E")   # тёмно-синяя шапка
ZEBRA_FILL = PatternFill("solid", fgColor="F2F5F9")    # светлая полоса зебры
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="2E3A4E")

_thin = Side(style="thin", color="D0D7E2")
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _auto_width(worksheet, products: list[Product]) -> None:
    # Ширина колонки = максимум длины содержимого (с запасом), но в разумных рамках.
    widths = [len(col) for col in COLUMNS]
    for product in products:
        values = [product.name, product.article,
                  "" if product.price is None else f"{product.price:.2f}",
                  product.category]
        for i, value in enumerate(values):
            widths[i] = max(widths[i], len(str(value)))
    for i, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = min(max(width + 3, 10), 60)


def write_report(products: list[Product], output_path: str | Path, title: str) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Прайс"
    worksheet.sheet_view.showGridLines = True  # сетка включена

    last_col = len(COLUMNS)

    # 1. Заголовок отчёта (объединяем по ширине таблицы).
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 22

    # 2. Шапка таблицы.
    header_row = 2
    for col, name in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=header_row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    # 3. Данные с зеброй.
    for i, product in enumerate(products):
        row = header_row + 1 + i
        values = [product.name, product.article, product.price, product.category]
        for col, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = CELL_BORDER
            if i % 2 == 1:  # чередование: подкрашиваем каждую вторую строку
                cell.fill = ZEBRA_FILL
        # Цена — как число с разделителем тысяч и двумя знаками.
        price_cell = worksheet.cell(row=row, column=3)
        price_cell.number_format = "# ##0.00"
        price_cell.alignment = Alignment(horizontal="right")

    # 4. Автоширина, автофильтр, закреплённая шапка.
    _auto_width(worksheet, products)
    end_row = header_row + len(products)
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{end_row}"
    worksheet.freeze_panes = f"A{header_row + 1}"

    workbook.save(output_path)
    return output_path
