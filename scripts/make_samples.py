"""Генерация примеров «кривых» прайсов для демонстрации.

Создаёт три файла в data/samples/ с типичными проблемами реальных прайсов:
  1. postavshik1_krepez.xlsx  — мусорная шапка + объединённые ячейки, крепёж.
  2. postavshik2_sbornyj.xlsx — другой порядок и названия колонок, разные категории,
                                строка-разделитель внутри данных.
  3. postavshik3_smeshannyj.csv — CSV с ; и мусором сверху.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

SAMPLES_DIR = Path(__file__).resolve().parent.parent / "data" / "samples"


def _make_krepez() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"

    # Мусор сверху + объединённые ячейки (шапка фирмы на всю ширину).
    ws.merge_cells("A1:D1")
    ws["A1"] = 'ООО «КрепёжТорг»'
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Прайс-лист от 01.07.2026 · тел. 8-800-100-20-30 · г. Новосибирск"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Строка 3 остаётся пустой — тоже мусор перед настоящей шапкой.
    # Настоящая шапка — строка 4 (пишем по адресам, т.к. строки 1-2 объединены).
    header = ["Наименование", "Артикул", "Цена, руб", "Остаток"]
    for col, value in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=value)

    rows = [
        ["Болт М8х40 DIN 933 оцинкованный", "KR-0801", 4.50, 1200],
        ["Гайка М8 DIN 934", "KR-0802", 1.80, 3500],
        ["Шайба плоская М8 DIN 125", "KR-0803", 0.60, 8000],
        ["Саморез по дереву 3.5х45", "KR-1145", 0.90, 15000],
        ["Дюбель распорный 6х40", "KR-2064", 1.20, 9000],
        ["Анкер клиновой 10х100", "KR-3110", 12.50, 750],
        ["Винт М6х20 с потайной головкой", "KR-0620", 3.10, 2000],
        ["Гвозди строительные 3.0х70", "KR-4070", 95.00, 400],
        ["Хомут червячный 20-32 мм", "KR-5032", 8.40, 1100],
        ["Заклёпка вытяжная 4х10", "KR-6410", 0.35, 20000],
    ]
    for i, row in enumerate(rows, start=5):
        for col, value in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=value)

    path = SAMPLES_DIR / "postavshik1_krepez.xlsx"
    wb.save(path)
    return path


def _make_sbornyj() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"  # типичное имя выгрузки из 1С

    # Мусор + объединённая шапка.
    ws.merge_cells("A1:D1")
    ws["A1"] = "База «СтройСнаб» — оптовый прайс"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A2:D2")
    ws["A2"] = "Цены действительны при заказе от 10 000 руб. Дата: 05.07.2026"

    # ДРУГОЙ порядок и названия колонок: код первым, цена — последней.
    header = ["Код", "Товар", "Кол-во на складе", "Стоимость, ₽"]
    for col, value in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=value)

    rows = [
        ["EL-1001", "Кабель ВВГнг 3х2.5", "1500 м", "72,50"],
        ["EL-1002", "Розетка встраиваемая с з/к", "320", "115,00"],
        ["EL-1003", "АвтоматC16 1P", "210", "245,00"],
        ["EL-1004", "Светильник LED 36Вт", "85", "890,00"],
        # Строка-разделитель внутри данных — есть текст, нет цены/кода.
        ["", "САНТЕХНИКА", "", ""],
        ["SN-2001", "Труба ПП 25 PN20", "800 м", "48,00"],
        ["SN-2002", "Кран шаровой 1/2\"", "460", "180,00"],
        ["SN-2003", "Смеситель для кухни", "70", "1 950,00"],
        ["SN-2004", "Сифон для раковины", "150", "210,00"],
        ["IN-3001", "Дрель ударная 750Вт", "40", "3 450,00"],
        ["IN-3002", "Шуруповёрт аккумуляторный 18В", "35", "5 990,00"],
        ["IN-3003", "Молоток слесарный 500 г", "120", "420,00"],
        ["KR-9001", "Саморезы по металлу 4.2х16 (уп. 200)", "600", "310,00"],
    ]
    for i, row in enumerate(rows, start=5):
        for col, value in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=value)

    path = SAMPLES_DIR / "postavshik2_sbornyj.xlsx"
    wb.save(path)
    return path


def _make_csv() -> Path:
    path = SAMPLES_DIR / "postavshik3_smeshannyj.csv"
    rows = [
        ["Поставщик: ИП Смирнов А.В."],
        ["Прайс на 07.07.2026, valid 14 дней"],
        [],
        ["Название", "Арт.", "Цена"],
        ["Выключатель одноклавишный белый", "V-100", "95,00"],
        ["Провод ПВС 2х1.5", "P-215", "38,50"],
        ["Труба гофрированная 20 мм", "T-020", "24,00"],
        ["Отвёртка крестовая PH2", "O-PH2", "150,00"],
        ["Уровень строительный 60 см", "U-060", "520,00"],
        ["Болт М10х50 оцинкованный", "B-1050", "6,80"],
        ["Герметик силиконовый 280 мл", "G-280", "230,00"],
        ["Лампа светодиодная E27 9Вт", "L-E27", "89,00"],
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerows(rows)
    return path


def main() -> None:
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    for maker in (_make_krepez, _make_sbornyj, _make_csv):
        path = maker()
        print(f"  ✓ {path.relative_to(SAMPLES_DIR.parent.parent)}")


if __name__ == "__main__":
    main()
